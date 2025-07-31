import os
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
import datetime

def generate_excel_report(Beginning, Span, report_type, df_dict):
        current_datetime = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        wb = Workbook()
        first_sheet = True

        for group_name, df in df_dict.items():
            # Create or get sheet
            if first_sheet:
                ws = wb.active
                ws.title = group_name
                first_sheet = False
            else:
                ws = wb.create_sheet(title=group_name)

            # Logo
            cwd = os.getcwd()
            image_path = os.path.join(cwd, 'assets/Logo.png')
            if os.path.exists(image_path):
                img = Image(image_path)
                img.width = 155
                img.height = 60
                ws.add_image(img, 'A1')

            # Metadata
            metadata = {
                "Beginning:": Beginning,
                "Span:": Span,
                "Report Generated:": current_datetime
            }

            for idx, (label, value) in enumerate(metadata.items(), start=1):
                ws[f'C{idx}'] = label
                ws[f'C{idx}'].font = Font(bold=True)
                ws[f'C{idx}'].alignment = Alignment(horizontal='right')
                ws[f'D{idx}'] = value
                ws[f'D{idx}'].alignment = Alignment(horizontal='left')

            # Header styling
            def apply_header_color(ws, start_row, header_color):
                for col_num, heading in enumerate(df.columns, 1):
                    cell = ws.cell(row=start_row, column=col_num, value=heading)
                    cell.font = Font(bold=True)
                    fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                    cell.fill = fill

            # Table data
            def insert_table(ws, start_row, header_color, df):
                apply_header_color(ws, start_row, header_color)
                for r_idx, row in df.iterrows():
                    for c_idx, value in enumerate(row, 1):
                        if pd.isna(value):
                            value = None
                        ws.cell(row=r_idx + start_row + 1, column=c_idx, value=value)

            # Auto-fit
            def autofit_columns(ws, df):
                for col_num, column in enumerate(df.columns, 1):
                    col_data = df[column].dropna().astype(str).tolist()
                    if not col_data:
                        continue
                    max_length = max(len(val) for val in col_data)
                    adjusted_width = max_length + 2
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = adjusted_width

            # Insert data
            insert_table(ws, 6, "D3D3D3", df)
            autofit_columns(ws, df)

        # Save the file
        file_name = f"{report_type}_{current_datetime}.xlsx"
        file_path = os.path.join(os.getcwd(), file_name)
        wb.save(file_path)
        return file_name